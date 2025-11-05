import openpyxl
import psycopg2
import datetime
import sys

def eksekusi_lineup():
    lineup_excel = openpyxl.load_workbook(r"lineup.xlsx", data_only=True).active
    data_lineup = lineup_excel.iter_rows(min_row=2, values_only=True)
    filtered_data = []
    for row in data_lineup:
    #     if any(cell is not None and str(cell).strip() != "" for cell in row):
    #         filtered_data.append(row)
        if row[1] not in [None, "#N/A"] and row[2] not in [None, "#N/A"] and row[3] not in [None, "#N/A"]:
            filtered_data.append(row)

    for row in filtered_data:
        no_unit = unit_cat(str(row[2]))
        fleet_category = fleet_cat(str(row[1]),str(row[2]))
        # row = list(row)
        # row[5] = row[5][:1] if row[5] and len(row[5]) == 1 else ""
        # row[6] = row[6][:1] if row[6] and len(row[6]) == 1 else ""
        cur_psql.execute("""insert into t_lineup (proddate, unit_no, jde_no, operator_name, crew, shift, location, download_date, download_time, fleet)
                values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) on conflict (proddate, unit_no, jde_no, shift) 
                --do nothing
                do update set operator_name = excluded.operator_name, 
                                crew = excluded.crew, 
                                location = excluded.location, 
                                download_date = excluded.download_date, 
                                download_time = excluded.download_time, 
                                fleet = excluded.fleet;
                """, (row[0], no_unit, row[3], row[4], row[5], row[6], row[7], tanggal, jam, fleet_category))
        print(row[0], row[1], no_unit, row[3])
    conn.commit()

def eksekusi_unit():
    unit_excel = openpyxl.load_workbook(r"\\10.30.241.2\lineup_mod\unit.xlsx").active
    data_unit = unit_excel.iter_rows(min_row=2, values_only=True)
    for row in data_unit:
        prefix = equipment_class(row[0])
        cur_psql.execute("insert into m_setup_unit (unit_no, egi, class) values (%s, %s, %s)", (row[0], row[1], prefix))
        conn.commit()

def eksekusi_opt():
    opt_excel = openpyxl.load_workbook(r"operator.xlsx", data_only=True).active
    data_opt = opt_excel.iter_rows(min_row=2, values_only=True)
    for row in data_opt:
        tipe_simper = row[4].replace(';',',').split(',')
        tipe_simper = [tipe for tipe in tipe_simper if tipe.strip() != '']
        for tipe in tipe_simper:           
            cur_psql.execute("""INSERT INTO m_operator (jde_no, operator_name, crew, simper_type, kategori)
                                VALUES (%s, %s, %s, %s, %s)""", (row[0], row[1], row[2], tipe, row[3]))

            print(f"Inserted : {row[0]}, {tipe}")
            conn.commit()

def del_lineup(tgl):
    cur_psql.execute(f"delete from t_lineup where proddate = '{tgl}'")
    cur_psql.execute("SELECT setval('t_lineup_id_seq', COALESCE((SELECT MAX(id) FROM t_lineup), 1))")
    conn.commit()

def bersihkan(qry):
    cur_psql.execute(qry)
    conn.commit()

def equipment_class(no_unit):
    kelas = no_unit[:2]
    if kelas in ['EX', 'DT', 'TD', 'GR', 'WT', 'CP']:
        return kelas
    elif kelas in ['HE']:
        return 'EX'
    return ""

def fleet_cat(fleet, spare):
    kelas = fleet[:2]
    if kelas in ['13', '14', '30', '40', '50', '31']:
        return 'EX' + fleet
    elif spare == 'SPARE':
        return 'SPARE'
    return fleet

def unit_cat(no_unit):
    tipe = no_unit[:2].upper()
    tipe2 = no_unit[:4]
    exca = ['5030','4027','4028','4029','3099','3073','3064','3059','1407','1382','1358','1322',
            '1431','1321','1377','1350','1343','1342','1378','1419','1334','1339','1434','1437']
    grader = ['4047','3012','4072','4073','3031','3047','3070','3076','3077']
    if tipe in ['11', '12']:
        return 'TD' + no_unit
    elif tipe2 in exca:
        return 'EX' + no_unit
    elif tipe == '31':
        return 'EX' + no_unit
    elif tipe2 in grader:
        return 'GR' + no_unit
    elif tipe in ['10', '20']:
        if tipe2 == '1053':
            return 'CM' + no_unit
        else:
            return 'WT' + no_unit
    elif tipe in ['21','23','24','25','27','28']:
        return 'DT' + no_unit
    return no_unit

# try:
waktu_awal = datetime.datetime.now()
conn = psycopg2.connect("information database credential in here")
print('Opened database postgreSQL succesfully')
cur_psql = conn.cursor()
tanggal = datetime.datetime.now().strftime('%Y-%m-%d')
jam = datetime.datetime.now().strftime('%H:%M:%S')

# bersihkan('truncate m_operator restart identity')
# eksekusi_opt()

# bersihkan('truncate m_setup_unit restart identity')
# eksekusi_unit()

# del_lineup(tanggal)
eksekusi_lineup()
waktu_akhir = datetime.datetime.now()
print(f'Waktu Eksekusi : {waktu_akhir - waktu_awal}')
# except Exception as e:
#     print(f'Error eksekusi : {e}')
#     sys.exit()